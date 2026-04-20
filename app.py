import streamlit as st
import pandas as pd
import re
import tempfile
import os
import zipfile
import xlsxwriter
import base64
import colorsys

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter
from datetime import datetime

st.set_page_config(page_title="THL TO SM", layout="centered")

# =========================
# CSS (GIỮ NGUYÊN)
# =========================
st.markdown("""
<style>
header {display: none !important;}
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
.block-container {padding-top: 0rem !important;}

.header {text-align: center; padding: 8px 0;}
.header h1 {color: #0284c7; margin: 0;}
.header p {color: #64748b; margin: 0;}

.card {background: white; padding: 20px; border-radius: 12px;}

.stButton>button {
    width: 100%;
    height: 42px;
    border-radius: 10px;
    background: linear-gradient(90deg, #0ea5e9, #22c55e);
    color: white;
}
</style>
""", unsafe_allow_html=True)

# =========================
# STATE
# =========================
if "uploader_key" not in st.session_state:
    st.session_state["uploader_key"] = 0

if "done" not in st.session_state:
    st.session_state["done"] = False

# =========================
# COLOR
# =========================
PASTEL_STRONG_DISTINCT = [
    "FFD6D6","FFE0EB","EBD6FF","D6E4FF","D6F5FF","D6FFF5",
    "D6FFD6","F0FFD6","FFF5D6","FFEBD6","FFDCD6","F5D6FF"
]

def generate_distinct_colors(n):
    colors = []
    base = PASTEL_STRONG_DISTINCT.copy()

    extra_needed = max(0, n - len(base))

    for i in range(extra_needed):
        h = (i * 0.17) % 1
        s = 0.25
        v = 0.95
        r, g, b = colorsys.hsv_to_rgb(h, s, v)
        colors.append('%02X%02X%02X' % (int(r*255), int(g*255), int(b*255)))

    return base + colors

# =========================
# AUTO COLUMN WIDTH FIX
# =========================
def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_length + 2

# =========================
# HEADER
# =========================
st.markdown("""
<div class="header">
    <h1>⚡ THL TO SM</h1>
    <p>Xử lý & đối soát Shipment nhanh chóng</p>
</div>
""", unsafe_allow_html=True)

# =========================
# UI
# =========================
with st.container():
    st.markdown('<div class="card">', unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "📂 Chọn 2 file Excel",
        type=["xlsx"],
        accept_multiple_files=True,
        key=f"uploader_{st.session_state['uploader_key']}"
    )

    ready = uploaded_files and len(uploaded_files) == 2

    if ready and not st.session_state["done"]:

        if st.button("🚀 Bắt đầu xử lý"):

            try:
                with st.spinner("⏳ Đang xử lý..."):

                    tmp_dir = tempfile.gettempdir()
                    path_tpn, path_book1 = None, None

                    for file in uploaded_files:
                        path = os.path.join(tmp_dir, file.name)
                        with open(path, "wb") as f:
                            f.write(file.read())

                        df_check = pd.read_excel(path, engine="calamine", dtype=str)

                        if any("Shipment Nbr" in str(c) for c in df_check.columns):
                            path_tpn = path
                        else:
                            path_book1 = path

                    save_path = os.path.join(tmp_dir, "TPN_KET_QUA.xlsx")
                    kehoach_path = os.path.join(tmp_dir, "TPN_KE_HOACH_XE.xlsx")

                    df2 = pd.read_excel(path_book1, engine="calamine", header=None, dtype=str)

                    group_list = []
                    for _, row in df2.iterrows():
                        nums = set()
                        text = "" if pd.isna(row.iloc[0]) else str(row.iloc[0])

                        for num in re.findall(r"\d+", text):
                            if len(num) == 3:
                                num = "0" + num
                            if len(num) == 4:
                                nums.add(num)

                        if nums:
                            group_list.append(nums)

                    df_tpn = pd.read_excel(path_tpn, engine="calamine", dtype=str)

                    wb = Workbook()
                    ws = wb.active

                    ws.append(list(df_tpn.columns))

                    for _, r in df_tpn.iterrows():
                        ws.append(list(r.values))

                    # =========================
                    # FIND COLUMNS
                    # =========================
                    col_index = None
                    date_col_index = None

                    for idx, c in enumerate(ws[1], start=1):
                        if "Shipment Nbr" in str(c.value):
                            col_index = idx
                        if "Shipment Date" in str(c.value):
                            date_col_index = idx

                    ketqua_numbers = set()

                    for i in range(2, ws.max_row + 1):
                        val = ws.cell(i, col_index).value
                        if val:
                            found = re.findall(r"\d+", str(val))
                            if found:
                                last = found[-1]
                                if len(last) == 3:
                                    last = "0" + last
                                if len(last) == 4:
                                    ketqua_numbers.add(last)

                    colors = generate_distinct_colors(len(group_list))
                    group_colors = {i: colors[i] for i in range(len(group_list))}

                    header_fill = PatternFill("solid", fgColor="000080")
                    header_font = Font(color="FFFFFF", bold=True)
                    bold_font = Font(bold=True)

                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font

                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            if cell.value:
                                cell.font = bold_font

                    count = 0

                    for i in range(2, ws.max_row + 1):
                        val = ws.cell(i, col_index).value
                        if val:
                            nums = set()
                            found = re.findall(r"\d+", str(val))

                            if found:
                                last = found[-1]
                                if len(last) == 3:
                                    last = "0" + last
                                if len(last) == 4:
                                    nums.add(last)

                            for idx, g in enumerate(group_list):
                                if nums & g:
                                    ws.cell(i, col_index).fill = PatternFill(
                                        fill_type="solid",
                                        fgColor=group_colors[idx]
                                    )
                                    count += 1
                                    break

                    ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

                    # =========================
                    # ✅ FIX SHIPMENT DATE (DD/MM/YYYY)
                    # =========================
                    if date_col_index:
                        for i in range(2, ws.max_row + 1):
                            cell = ws.cell(i, date_col_index)

                            if cell.value:
                                try:
                                    if isinstance(cell.value, str):
                                        cell.value = datetime.strptime(cell.value[:10], "%Y-%m-%d")

                                    cell.number_format = "DD/MM/YYYY"
                                except:
                                    pass

                    auto_adjust_column_width(ws)

                    wb.save(save_path)
                    wb.close()

                    workbook = xlsxwriter.Workbook(kehoach_path)
                    worksheet = workbook.add_worksheet()

                    red = workbook.add_format({'font_color': 'red'})
                    normal = workbook.add_format({})

                    col_width = 0

                    for r, row in df2.iterrows():
                        text = "" if pd.isna(row.iloc[0]) else str(row.iloc[0])
                        col_width = max(col_width, len(text))

                        parts = []
                        last = 0

                        for m in re.finditer(r"\d+", text):
                            num = m.group()
                            start, end = m.span()
                            check = "0"+num if len(num) == 3 else num

                            if start > last:
                                parts += [normal, text[last:start]]

                            parts += [red if check in ketqua_numbers else normal, num]
                            last = end

                        if last < len(text):
                            parts += [normal, text[last:]]

                        try:
                            worksheet.write_rich_string(r, 0, *parts)
                        except:
                            worksheet.write(r, 0, text)

                    worksheet.set_column(0, 0, col_width + 3)
                    workbook.close()

                    zip_path = os.path.join(tmp_dir, "TPN_COMPLETE.zip")
                    with zipfile.ZipFile(zip_path, "w") as z:
                        z.write(save_path, "TPN_KET_QUA.xlsx")
                        z.write(kehoach_path, "TPN_KE_HOACH_XE.xlsx")

                    with open(zip_path, "rb") as f:
                        zip_data = f.read()

                st.success(f"✅ COMPLETE !!! Matched: {count}")

                b64 = base64.b64encode(zip_data).decode()
                st.components.v1.html(f"""
                    <a id="dl" href="data:application/zip;base64,{b64}" download="THL TO SM.zip"></a>
                    <script>document.getElementById('dl').click();</script>
                """, height=0)

                st.session_state["done"] = True

            except Exception as e:
                st.error(f"❌ Lỗi: {e}")

    if st.session_state["done"]:
        if st.button("🔄 Xử lý file mới"):
            st.session_state["uploader_key"] += 1
            st.session_state["done"] = False
            st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)
